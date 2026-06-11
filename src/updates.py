"""
Updates feed: exporta la hoja "Updates Log" del Excel a docs/data/updates.json.

La hoja la mantiene la skill update-watchlist-ticker (una fila por update:
Ticker | Fecha | Campos cambiados | Fuente | Resumen). El frontend la consume
para el panel "últimas publicaciones" y la ficha (drawer) de cada empresa.
"""
from __future__ import annotations

import json
import logging
from pathlib import Path

log = logging.getLogger(__name__)

SHEET = "Updates Log"


def build_updates(xlsx_path, output_path="docs/data/updates.json") -> dict:
    """Lee la hoja Updates Log y escribe updates.json. Devuelve meta-stats."""
    import openpyxl

    xlsx_path = Path(xlsx_path)
    output_path = Path(output_path)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        log.warning("Hoja '%s' no encontrada en %s — skip updates feed", SHEET, xlsx_path)
        return {"n_updates": 0, "skipped": True}

    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {"n_updates": 0}

    # Fila 1 = headers: Ticker | Fecha | Campos cambiados | Fuente | Resumen
    entries = []
    for r in rows[1:]:
        if r is None or r[0] is None:
            continue
        ticker = str(r[0]).strip()
        fecha = str(r[1]).strip()[:10] if len(r) > 1 and r[1] is not None else None
        campos = str(r[2]).strip() if len(r) > 2 and r[2] is not None else ""
        fuente = str(r[3]).strip() if len(r) > 3 and r[3] is not None else ""
        resumen = str(r[4]).strip() if len(r) > 4 and r[4] is not None else ""
        entries.append({
            "ticker": ticker,
            "date": fecha,
            "fields_changed": campos,
            "source": fuente,
            "summary": resumen,
        })

    # Más recientes primero (fecha desc)
    entries.sort(key=lambda e: (e["date"] or "", ), reverse=True)

    payload = {
        "meta": {
            "n_updates": len(entries),
            "source_sheet": SHEET,
            "source_file": xlsx_path.name,
        },
        "updates": entries,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    log.info("updates.json -> %s (%d updates)", output_path, len(entries))
    return {"n_updates": len(entries)}


def main():
    import argparse
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    p = argparse.ArgumentParser(description="Exporta Updates Log del Excel a updates.json")
    p.add_argument("--xlsx", default="data/raw/watchlist_ratings.xlsx")
    p.add_argument("--out", default="docs/data/updates.json")
    args = p.parse_args()
    stats = build_updates(args.xlsx, output_path=args.out)
    print(stats)


if __name__ == "__main__":
    main()
