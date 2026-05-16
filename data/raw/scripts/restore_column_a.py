"""
restore_column_a.py
───────────────────
Copia el contenido de la columna A (desde A4 hasta la última fila con
contenido en col A) desde el backup más reciente hacia el Excel principal.

Se llama SIEMPRE al final de cualquier actualización automática o manual
para garantizar que los valores/fórmulas de la columna A no queden
corrompidos por openpyxl.

Uso:
    python restore_column_a.py                        # usa rutas por defecto
    python restore_column_a.py --excel /ruta/otro.xlsx
    python restore_column_a.py --backup /ruta/backup.xlsx  # fuerza un backup concreto
    python restore_column_a.py --dry-run              # sólo muestra qué haría

Salida JSON (stdout):
    {
      "status": "ok",
      "rows_restored": 61,
      "first_row": 4,
      "last_row": 64,
      "backup_used": "watchlist_ratings.backup_20260430_1537.xlsx",
      "excel_saved": "watchlist_ratings.xlsx"
    }
"""

import argparse
import glob
import json
import os
import sys
from datetime import datetime

# ── rutas por defecto ─────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
DATA_DIR    = os.path.dirname(SCRIPT_DIR)
EXCEL_FILE  = os.path.join(DATA_DIR, "watchlist_ratings.xlsx")
BACKUP_GLOB = os.path.join(DATA_DIR, "watchlist_ratings.backup_*.xlsx")

SHEET_NAME  = "Watchlist Ratings"
COL_A       = 1
HEADER_ROW  = 3
DATA_START  = 4


# ── helpers ───────────────────────────────────────────────────────────────────
def latest_backup(data_dir: str) -> str | None:
    """Devuelve la ruta del backup más reciente, o None si no hay ninguno."""
    pattern = os.path.join(data_dir, "watchlist_ratings.backup_*.xlsx")
    backups = sorted(glob.glob(pattern))
    return backups[-1] if backups else None


def find_last_content_row(ws, col: int, start_row: int) -> int:
    """Última fila en la columna `col` (1-based) que tenga valor no nulo."""
    last = start_row - 1
    for row in ws.iter_rows(min_row=start_row, min_col=col, max_col=col, values_only=False):
        cell = row[0]
        if cell.value is not None:
            last = cell.row
    return last


def copy_cell(src_cell, dst_cell):
    """
    Copia valor, tipo de dato y formato numérico de src a dst.
    Preserva el tipo 'e' (error) tal como está almacenado.
    """
    dst_cell.value      = src_cell.value
    dst_cell.data_type  = src_cell.data_type
    if src_cell.number_format:
        dst_cell.number_format = src_cell.number_format


# ── lógica principal ──────────────────────────────────────────────────────────
def restore_column_a(
    excel_path: str,
    backup_path: str | None = None,
    dry_run: bool = False,
) -> dict:
    import openpyxl

    # ── resolver backup ──────────────────────────────────────────────────────
    if backup_path is None:
        backup_path = latest_backup(os.path.dirname(excel_path))

    if backup_path is None:
        return {
            "status": "error",
            "message": "No se encontró ningún backup en el directorio.",
        }

    if not os.path.exists(backup_path):
        return {
            "status": "error",
            "message": f"Backup no encontrado: {backup_path}",
        }

    # ── cargar backup (fuente) ───────────────────────────────────────────────
    try:
        wb_bk = openpyxl.load_workbook(backup_path, data_only=False)
    except Exception as exc:
        return {"status": "error", "message": f"No se pudo abrir el backup: {exc}"}

    if SHEET_NAME not in wb_bk.sheetnames:
        return {
            "status": "error",
            "message": f"Hoja '{SHEET_NAME}' no encontrada en el backup.",
        }
    ws_bk = wb_bk[SHEET_NAME]

    # ── determinar rango ─────────────────────────────────────────────────────
    last_row_bk = find_last_content_row(ws_bk, COL_A, DATA_START)

    if last_row_bk < DATA_START:
        return {
            "status": "warning",
            "message": "No hay contenido en col A del backup a partir de la fila 4.",
        }

    if dry_run:
        wb_bk.close()
        return {
            "status": "dry-run",
            "rows_would_restore": last_row_bk - DATA_START + 1,
            "first_row": DATA_START,
            "last_row": last_row_bk,
            "backup_used": os.path.basename(backup_path),
            "excel_target": os.path.basename(excel_path),
        }

    # ── cargar Excel principal (destino) ─────────────────────────────────────
    try:
        wb_main = openpyxl.load_workbook(excel_path, data_only=False)
    except Exception as exc:
        wb_bk.close()
        return {"status": "error", "message": f"No se pudo abrir el Excel principal: {exc}"}

    if SHEET_NAME not in wb_main.sheetnames:
        wb_bk.close()
        wb_main.close()
        return {
            "status": "error",
            "message": f"Hoja '{SHEET_NAME}' no encontrada en el Excel principal.",
        }
    ws_main = wb_main[SHEET_NAME]

    # ── copiar A[DATA_START:last_row] ─────────────────────────────────────────
    rows_restored = 0
    for r in range(DATA_START, last_row_bk + 1):
        src = ws_bk.cell(row=r, column=COL_A)
        dst = ws_main.cell(row=r, column=COL_A)
        copy_cell(src, dst)
        rows_restored += 1

    # ── guardar ───────────────────────────────────────────────────────────────
    try:
        wb_main.save(excel_path)
    except PermissionError:
        wb_bk.close()
        wb_main.close()
        return {
            "status": "error",
            "message": (
                f"No se pudo guardar '{excel_path}'. "
                "¿Está abierto en Excel? Ciérralo e inténtalo de nuevo."
            ),
        }

    wb_bk.close()
    wb_main.close()

    return {
        "status": "ok",
        "rows_restored": rows_restored,
        "first_row": DATA_START,
        "last_row": last_row_bk,
        "backup_used": os.path.basename(backup_path),
        "excel_saved": os.path.basename(excel_path),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


# ── CLI ───────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Restaura la columna A del Excel desde el backup más reciente."
    )
    parser.add_argument(
        "--excel",  default=EXCEL_FILE,
        help=f"Ruta al Excel principal (default: {EXCEL_FILE})"
    )
    parser.add_argument(
        "--backup", default=None,
        help="Ruta a un backup concreto (default: el más reciente en data/raw)"
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Muestra qué haría sin modificar ningún archivo."
    )
    args = parser.parse_args()

    result = restore_column_a(
        excel_path=args.excel,
        backup_path=args.backup,
        dry_run=args.dry_run,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))
    sys.exit(0 if result.get("status") in ("ok", "dry-run", "warning") else 1)


if __name__ == "__main__":
    main()
