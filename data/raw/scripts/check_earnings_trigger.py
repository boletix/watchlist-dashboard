"""
check_earnings_trigger.py
─────────────────────────
Escanea watchlist_ratings.xlsx y devuelve los tickers cuya columna
"Últimos resultados" (col BO) + N días == hoy (defecto: 5).

Uso:
    python check_earnings_trigger.py                  # días = 5 (default)
    python check_earnings_trigger.py --days 5         # explícito
    python check_earnings_trigger.py --date 2026-04-30 # simular otra fecha

Salida JSON (stdout):
    {
      "run_date": "2026-04-30",
      "days_offset": 5,
      "due": [
        {"ticker": "MSFT", "row": 5, "last_results_date": "2026-04-25"},
        ...
      ],
      "skipped": [
        {"ticker": "JDG", "row": 26, "reason": "N/D"},
        ...
      ]
    }
"""

import argparse
import json
import os
import sys
from datetime import date, datetime, timedelta
import glob

# ── rutas ────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
DATA_DIR     = os.path.dirname(SCRIPT_DIR)          # …/data/raw
EXCEL_FILE   = os.path.join(DATA_DIR, "watchlist_ratings.xlsx")
BACKUP_GLOB  = os.path.join(DATA_DIR, "watchlist_ratings.backup_*.xlsx")

# columnas (1-based, como openpyxl)
COL_TICKER           = 2   # B
COL_ULTIMOS_RESULT   = 67  # BO  → "Últimos resultados"
COL_PROXIMOS_RESULT  = 68  # BP  → "Próximos resultados"

HEADER_ROW  = 3
DATA_START  = 4
DATE_FMTS   = ["%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"]


# ── helpers ──────────────────────────────────────────────────────────────────
def parse_date(raw) -> date | None:
    """Intenta parsear la celda como fecha. Devuelve None si no puede."""
    if raw is None:
        return None
    if isinstance(raw, (datetime,)):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    if not s or s.upper() in ("N/D", "N/A", "-", ""):
        return None
    for fmt in DATE_FMTS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def best_excel_source() -> str:
    """
    Devuelve la ruta al Excel a leer: el principal si está bien formado,
    o el backup más reciente si el principal está bloqueado/corrupto.
    """
    if os.path.exists(EXCEL_FILE):
        try:
            import openpyxl
            openpyxl.load_workbook(EXCEL_FILE, read_only=True).close()
            return EXCEL_FILE
        except Exception:
            pass

    backups = sorted(glob.glob(BACKUP_GLOB))
    if backups:
        return backups[-1]
    raise FileNotFoundError(
        f"No se encontró Excel válido en {DATA_DIR}. "
        "Asegúrate de que watchlist_ratings.xlsx existe y no está dañado."
    )


# ── lógica principal ─────────────────────────────────────────────────────────
def find_due_tickers(run_date: date, days_offset: int = 5) -> dict:
    import openpyxl

    source_path = best_excel_source()
    wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)

    if "Watchlist Ratings" not in wb.sheetnames:
        raise ValueError("No se encontró la hoja 'Watchlist Ratings' en el Excel.")

    ws = wb["Watchlist Ratings"]

    due     = []
    skipped = []
    target_date = run_date - timedelta(days=days_offset)  # fecha BO que buscamos

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=DATA_START, values_only=True), start=DATA_START
    ):
        ticker = row[COL_TICKER - 1] if len(row) >= COL_TICKER else None
        if not ticker:
            continue  # fila vacía → skip silencioso

        raw_bo = row[COL_ULTIMOS_RESULT - 1] if len(row) >= COL_ULTIMOS_RESULT else None
        last_results = parse_date(raw_bo)

        if last_results is None:
            skipped.append({
                "ticker": str(ticker),
                "row": row_idx,
                "reason": str(raw_bo) if raw_bo is not None else "vacío"
            })
            continue

        if last_results == target_date:
            raw_bp = row[COL_PROXIMOS_RESULT - 1] if len(row) >= COL_PROXIMOS_RESULT else None
            due.append({
                "ticker": str(ticker),
                "row": row_idx,
                "last_results_date": last_results.strftime("%Y-%m-%d"),
                "next_results_date": parse_date(raw_bp).strftime("%Y-%m-%d")
                    if parse_date(raw_bp) else str(raw_bp)
            })

    wb.close()

    return {
        "run_date":   run_date.strftime("%Y-%m-%d"),
        "days_offset": days_offset,
        "source_file": os.path.basename(source_path),
        "due":     due,
        "skipped": skipped,
    }


# ── CLI ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Detecta tickers con resultados hace exactamente N días."
    )
    parser.add_argument(
        "--days", type=int, default=5,
        help="Días desde los resultados para disparar la actualización (default: 5)"
    )
    parser.add_argument(
        "--date", type=str, default=None,
        help="Simula otra fecha de ejecución (YYYY-MM-DD). Default: hoy."
    )
    args = parser.parse_args()

    run_date = date.today()
    if args.date:
        try:
            run_date = datetime.strptime(args.date, "%Y-%m-%d").date()
        except ValueError:
            print(f"ERROR: --date debe tener formato YYYY-MM-DD, recibido: {args.date}", file=sys.stderr)
            sys.exit(1)

    try:
        result = find_due_tickers(run_date=run_date, days_offset=args.days)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except Exception as exc:
        print(json.dumps({"error": str(exc)}), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
