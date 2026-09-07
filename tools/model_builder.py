"""
Generador del modelo financiero de 12 hojas, estandar Prysmian.

Hojas: Dashboard, Assumptions, Revenue_Build, COGS_GP, PL, Working_Capital,
       Cash_Flow, Balance, Capital_Alloc, Valuation, Sensitivities, Sources

Convenciones (identicas a Prysmian_Model_2019-2030.xlsx):
  - Assumptions!B4  = escenario activo (1 conservador, 2 base, 3 alcista)
  - Assumptions!B8  = precio por accion  <- ACTUALIZAR antes de usar la valoracion
  - azul = input · verde = historico reportado · negro = formula · amarillo = supuesto
  - negativos en rojo entre parentesis, todo en millones de la moneda de REPORTE

La hoja Valuation calcula LOS TRES ESCENARIOS A LA VEZ (no depende de B4): lee las
columnas C/D/E de Assumptions. B4 solo mueve el modelo operativo.

TIR: los tres escenarios se anualizan SIEMPRE a 5 anios. Fue el bug de agosto de 2026
(alcista a 4 anios, bajista a 5) y no puede volver.

Las filas de cada hoja son FIJAS y estan declaradas como constantes abajo, para que las
hojas puedan referenciarse entre si sin depender del orden de construccion.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- estilo

BLUE = Font(color="0000CC", name="Calibri", size=10)
GREEN = Font(color="006600", name="Calibri", size=10)
BLACK = Font(color="000000", name="Calibri", size=10)
BOLD = Font(bold=True, name="Calibri", size=10)
TITLE = Font(bold=True, size=14, name="Calibri", color="1F3864")
H1 = Font(bold=True, size=11, name="Calibri", color="FFFFFF")
SMALL = Font(italic=True, size=8, name="Calibri", color="666666")

FILL_HDR = PatternFill("solid", fgColor="1F3864")
FILL_KEY = PatternFill("solid", fgColor="FFF2CC")
FILL_BAND = PatternFill("solid", fgColor="F2F2F2")
FILL_PROJ = PatternFill("solid", fgColor="FDF7E3")
FILL_OK = PatternFill("solid", fgColor="E2EFDA")

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM = '#,##0;[Red](#,##0)'
NUM1 = '#,##0.0;[Red](#,##0.0)'
NUM2 = '#,##0.00;[Red](#,##0.00)'
PCT = '0.0%;[Red](0.0%)'
MULT = '0.0"x"'
RATIO = '0.00":1"'

C0 = 3  # primera columna de anios = C

# --------------------------------------------------- mapa de filas (FIJO)

# Assumptions
A_SCEN, A_PRICE, A_SHARES, A_MCAP = 4, 8, 9, 10
A_DEBT, A_CASH, A_ND, A_EV, A_FCF_LTM, A_EVFCF, A_FX = 11, 12, 13, 14, 15, 16, 17
A_HDR = 18            # cabecera de la tabla de escenarios
A_FIRST = 19          # primer supuesto

ASSUMPTIONS = [
    # (etiqueta, unidad, formato)
    ("Crecimiento organico anual", "%", PCT),
    ("Crecimiento por adquisiciones anual", "%", PCT),
    ("Margen bruto de salida", "%", PCT),
    ("Opex de salida (% ingresos)", "%", PCT),
    ("D&A (% ingresos)", "%", PCT),
    ("Capex (% ingresos)", "%", PCT),
    ("Circulante neto (% ingresos)", "%", PCT),
    ("Tasa fiscal efectiva", "%", PCT),
    ("Coste medio de la deuda", "%", PCT),
    ("SBC (% ingresos)", "%", PCT),
    ("M&A (% del FCF)", "%", PCT),
    ("Dividendos (% del FCF)", "%", PCT),
    ("Recompras (% del FCF)", "%", PCT),
    ("Margen FCF de salida", "%", PCT),
    ("Multiplo de salida EV/FCF", "x", MULT),
    ("Capex de MANTENIMIENTO (% ingresos)", "%", PCT),
    ("I+D de MANTENIMIENTO (% ingresos)", "%", PCT),
]
A_ROW = {label: A_FIRST + i for i, (label, _, _) in enumerate(ASSUMPTIONS)}
A_LAST = A_FIRST + len(ASSUMPTIONS) - 1

# Revenue_Build: cabecera en 3, datos desde 5
RB_REV, RB_ORG, RB_MA, RB_TOT = 5, 6, 7, 8
RB_CHK = 10

# COGS_GP
CG_REV, CG_GM, CG_GP, CG_COGS = 5, 6, 7, 8

# PL
PL_REV, PL_GP, PL_OPEX, PL_EBIT, PL_MEBIT = 5, 6, 7, 8, 9
PL_DA, PL_EBITDA, PL_MEBITDA, PL_INT, PL_PBT, PL_TAX, PL_NI = 10, 11, 12, 13, 14, 15, 16

# Working_Capital: cabecera en 4, datos desde 6
WC_REC, WC_INV, WC_PAY, WC_NET, WC_PCT, WC_DELTA = 6, 7, 8, 9, 10, 11
WC_CHK = 14

# Cash_Flow
CF_EBITDA, CF_INT, CF_TAX, CF_DWC, CF_OTHER = 5, 6, 7, 8, 9
CF_OCF, CF_CAPEX, CF_FCF, CF_MARGIN = 10, 11, 12, 13
CF_SBC, CF_FCF_SBC = 15, 16
# Bloque de FCF NORMALIZADO (filas anadidas debajo; no mueven a las de arriba)
CF_NORM_HDR, CF_MAINT_CAPEX, CF_GROWTH_CAPEX = 18, 19, 20
CF_RND, CF_RND_STEADY, CF_RND_GROWTH, CF_FCF_NORM, CF_NORM_WARN = 21, 22, 23, 24, 25

# Balance
BS_CASH, BS_DEBT, BS_ND, BS_LEV, BS_EQ, BS_GW, BS_ASSETS, BS_CE, BS_ROIC = 5, 6, 7, 8, 9, 10, 11, 12, 13

# Capital_Alloc: cabecera en 4, datos desde 6
CA_FCF, CA_MA, CA_DIV, CA_BB, CA_RET, CA_SHRET = 6, 7, 8, 9, 10, 11
CA_ROIIC = 14

# Valuation
V_REV, V_MFCF, V_FCF, V_SBC, V_FCFADJ, V_FCF_XL = 6, 7, 8, 9, 10, 11
V_EVFCF5 = 12          # <- la metrica de cabecera: EV de HOY / FCF del anio 5
V_MULT, V_EV, V_ND, V_EQ, V_SH, V_PS, V_PX, V_UP, V_IRR = 13, 14, 15, 16, 17, 18, 19, 20, 21
V_ASYM, V_MIN, V_VERD = 23, 24, 25
V_BEARCHK = 28
V_NORM_HDR, V_FCFNORM, V_EVFCFNORM, V_NORMCHK = 33, 34, 35, 36
V_RECON_HDR, V_RECON_OP, V_RECON_CAL, V_RECON_GAP = 38, 39, 40, 41


def _cell(ws, row, col, value, font=None, fmt=None, fill=None, align=None, border=False,
          wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    if font:
        c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if align or wrap:
        c.alignment = Alignment(horizontal=align, wrap_text=wrap, vertical="top")
    if border:
        c.border = BOX
    return c


def _section(ws, row, text, width):
    for col in range(1, width + 1):
        _cell(ws, row, col, text if col == 1 else None, font=H1, fill=FILL_HDR)
    return row + 1


class ModelBuilder:
    def __init__(self, cd, cfg):
        self.cd = cd
        self.cfg = cfg
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

        hist = list(cd.years[-5:]) if cd.years else []
        last = hist[-1] if hist else 2025
        self.hist_years = hist
        self.proj_years = [last + i for i in range(1, 6)]
        self.years = hist + self.proj_years
        self.n_hist = len(hist)
        self.n_proj = 5
        self.term = self.years[-1]
        self.lh = get_column_letter(C0 + self.n_hist - 1) if self.n_hist else "C"
        self.tc = get_column_letter(C0 + len(self.years) - 1)

    # ------------------------------------------------------------ helpers

    def col(self, year):
        return C0 + self.years.index(year)

    def is_proj(self, year):
        return year in self.proj_years

    def A(self, label):
        return "Assumptions!$F$%d" % A_ROW[label]

    def AS(self, label, scen):
        return "Assumptions!$%s$%d" % ("CDE"[scen - 1], A_ROW[label])

    def year_header(self, ws, row):
        _cell(ws, row, 1, "Ejercicio", font=BOLD)
        _cell(ws, row, 2, "Unidad", font=BOLD)
        for y in self.years:
            c = _cell(ws, row, self.col(y), y, font=BOLD, align="center")
            if self.is_proj(y):
                c.fill = FILL_PROJ
        _cell(ws, row + 1, 1, "  verde = reportado · negro = formula · amarillo = supuesto",
              font=SMALL)
        for c, w in (("A", 44), ("B", 10)):
            ws.column_dimensions[c].width = w
        for y in self.years:
            ws.column_dimensions[get_column_letter(self.col(y))].width = 13
        return row + 2

    def line(self, ws, row, label, unit, hist_vals=None, proj=None, hist=None,
             fmt=NUM, bold=False, fill_row=None):
        f = BOLD if bold else BLACK
        _cell(ws, row, 1, label, font=f)
        _cell(ws, row, 2, unit, font=SMALL)
        for y in self.years:
            col = self.col(y)
            gl = get_column_letter(col)
            if self.is_proj(y):
                v = proj(y, gl, col) if proj else None
                cc = _cell(ws, row, col, v, font=f, fmt=fmt)
                cc.fill = fill_row or FILL_PROJ
            else:
                if hist:
                    cc = _cell(ws, row, col, hist(y, gl, col), font=f, fmt=fmt)
                else:
                    cc = _cell(ws, row, col, (hist_vals or {}).get(y), font=GREEN, fmt=fmt)
                if fill_row:
                    cc.fill = fill_row
        return row + 1

    def neg(self, metric):
        return {y: -abs(v) for y, v in self.cd.hist.get(metric, {}).items() if v is not None}

    # ------------------------------------------------------------ build

    def build(self):
        self.sh_assumptions()
        self.sh_revenue()
        self.sh_cogs()
        self.sh_pl()
        self.sh_wc()
        self.sh_cf()
        self.sh_balance()
        self.sh_capalloc()
        self.sh_valuation()
        self.sh_sensitivities()
        self.sh_sources()
        self.sh_dashboard()
        order = ["Dashboard", "Assumptions", "Revenue_Build", "COGS_GP", "PL",
                 "Working_Capital", "Cash_Flow", "Balance", "Capital_Alloc",
                 "Valuation", "Sensitivities", "Sources"]
        self.wb._sheets = [self.wb[n] for n in order]
        return self.wb

    # --- Assumptions -------------------------------------------------

    def sh_assumptions(self):
        cd, cfg = self.cd, self.cfg
        ws = self.wb.create_sheet("Assumptions")
        ws.column_dimensions["A"].width = 46
        ws.column_dimensions["B"].width = 14
        for c in "CDEF":
            ws.column_dimensions[c].width = 15

        _cell(ws, 1, 1, "%s (%s) — Supuestos" % (cd.name, cd.ticker), font=TITLE)
        _cell(ws, 2, 1, "Millones de %s. Fuente del historico en la hoja Sources." % cd.currency,
              font=SMALL)

        _section(ws, 3, "PANEL DE CONTROL", 6)
        _cell(ws, A_SCEN, 1, "Escenario activo (1=conservador, 2=base, 3=alcista)", font=BOLD)
        _cell(ws, A_SCEN, 2, 2, font=BLUE, fmt=NUM, fill=FILL_KEY, border=True)
        _cell(ws, 5, 1, "Escenario", font=BLACK)
        _cell(ws, 5, 2, '=CHOOSE($B$4,"Conservador","Base","Alcista")', font=BLACK)
        _cell(ws, 6, 1, "  Cambia B4 y recalcula el modelo operativo entero. "
                        "Valuation muestra los tres a la vez.", font=SMALL)

        _section(ws, 7, "MERCADO — ACTUALIZAR EL PRECIO ANTES DE VALORAR", 6)
        mkt = [
            (A_PRICE, "Precio por accion", cfg["price"], NUM2, BLUE),
            (A_SHARES, "Acciones en circulacion (M)", cfg["shares"], NUM1, BLUE),
            (A_MCAP, "Capitalizacion (en moneda de reporte)",
             "=$B$%d/$B$%d*$B$%d" % (A_PRICE, A_FX, A_SHARES), NUM, BLACK),
            (A_DEBT, "Deuda total", cfg["debt"], NUM, BLUE),
            (A_CASH, "Caja y equivalentes", cfg["cash"], NUM, BLUE),
            (A_ND, "Deuda neta", "=$B$%d-$B$%d" % (A_DEBT, A_CASH), NUM, BLACK),
            (A_EV, "Enterprise value", "=$B$%d+$B$%d" % (A_MCAP, A_ND), NUM, BLACK),
            (A_FCF_LTM, "FCF LTM", cfg["fcf_ltm"], NUM, BLUE),
            (A_EVFCF, "EV / FCF LTM", "=IFERROR($B$%d/$B$%d,\"n/d\")" % (A_EV, A_FCF_LTM),
             MULT, BLACK),
            (A_FX, "Tipo de cambio  %s -> %s" % (cd.currency, cfg.get("quote_ccy", cd.currency)),
             cfg.get("fx", 1.0), NUM2, BLUE),
        ]
        for row, label, val, fmt, font in mkt:
            _cell(ws, row, 1, label, font=BOLD if label in ("Enterprise value", "EV / FCF LTM")
                  else BLACK)
            _cell(ws, row, 2, val, font=font, fmt=fmt, border=True,
                  fill=FILL_KEY if font is BLUE else None)
        _cell(ws, A_FX, 3, "reporta en %s, cotiza en %s. 1 %s = %s %s"
              % (cd.currency, cfg.get("quote_ccy", cd.currency), cd.currency,
                 cfg.get("fx", 1.0), cfg.get("quote_ccy", cd.currency)), font=SMALL)

        _section(ws, A_HDR, "SUPUESTOS POR ESCENARIO", 6)
        for j, t in enumerate(("Conservador", "Base", "Alcista", "ACTIVO")):
            _cell(ws, A_HDR, 3 + j, t, font=H1, fill=FILL_HDR, align="center")

        vals = cfg["assumptions"]
        for label, unit, fmt in ASSUMPTIONS:
            r = A_ROW[label]
            _cell(ws, r, 1, label, font=BLACK)
            _cell(ws, r, 2, unit, font=SMALL)
            trio = vals.get(label, (0, 0, 0))
            for j, v in enumerate(trio):
                _cell(ws, r, 3 + j, v, font=BLUE, fmt=fmt, fill=FILL_KEY, border=True)
            _cell(ws, r, 6, "=INDEX($C%d:$E%d,1,$B$4)" % (r, r), font=BOLD, fmt=fmt,
                  fill=FILL_OK, border=True)

        r = A_LAST + 2
        _cell(ws, r, 1, "Notas del escenario", font=BOLD)
        for i, n in enumerate(cfg.get("scenario_notes", [])):
            _cell(ws, r + 1 + i, 1, "· " + n, font=SMALL)
        return ws

    # --- Revenue_Build ------------------------------------------------

    def sh_revenue(self):
        cd = self.cd
        ws = self.wb.create_sheet("Revenue_Build")
        _cell(ws, 1, 1, "%s — Construccion de ingresos" % cd.name, font=TITLE)
        self.year_header(ws, 3)

        self.line(ws, RB_REV, "Ingresos", "M", cd.hist.get("revenue", {}),
                  proj=lambda y, gl, c: "=%s%d*(1+%s%d)" % (
                      get_column_letter(c - 1), RB_REV, gl, RB_TOT), bold=True)
        self.line(ws, RB_ORG, "  Crecimiento organico", "%", {},
                  proj=lambda y, gl, c: "=%s" % self.A("Crecimiento organico anual"), fmt=PCT)
        self.line(ws, RB_MA, "  Crecimiento por adquisiciones", "%", {},
                  proj=lambda y, gl, c: "=%s" % self.A("Crecimiento por adquisiciones anual"),
                  fmt=PCT)
        self.line(ws, RB_TOT, "  Crecimiento total", "%",
                  hist=lambda y, gl, c: ("=IFERROR(%s%d/%s%d-1,\"\")" % (
                      gl, RB_REV, get_column_letter(c - 1), RB_REV)) if c > C0 else None,
                  proj=lambda y, gl, c: "=%s%d+%s%d" % (gl, RB_ORG, gl, RB_MA),
                  fmt=PCT, bold=True)

        _cell(ws, RB_CHK, 1, "Comprobacion contra el Excel de la watchlist", font=BOLD)
        _cell(ws, RB_CHK + 1, 1, "Ingresos LTM segun watchlist_ratings.xlsx", font=BLACK)
        _cell(ws, RB_CHK + 1, 3, cd.wl.get("revenue_ltm"), font=BLUE, fmt=NUM)
        _cell(ws, RB_CHK + 2, 1, "Ingresos LTM segun yfinance (4 trimestres)", font=BLACK)
        _cell(ws, RB_CHK + 2, 3, cd.ttm.get("revenue"), font=GREEN, fmt=NUM)
        _cell(ws, RB_CHK + 3, 1, "Diferencia (si es grande, revisar el Excel)", font=BOLD)
        _cell(ws, RB_CHK + 3, 3, "=C%d-C%d" % (RB_CHK + 1, RB_CHK + 2), font=BOLD, fmt=NUM)
        return ws

    # --- COGS_GP -------------------------------------------------------

    def sh_cogs(self):
        cd = self.cd
        ws = self.wb.create_sheet("COGS_GP")
        _cell(ws, 1, 1, "%s — Coste de ventas y margen bruto" % cd.name, font=TITLE)
        self.year_header(ws, 3)
        lh, n = self.lh, self.n_proj

        self.line(ws, CG_REV, "Ingresos", "M",
                  hist=lambda y, gl, c: "=Revenue_Build!%s%d" % (gl, RB_REV),
                  proj=lambda y, gl, c: "=Revenue_Build!%s%d" % (gl, RB_REV))
        self.line(ws, CG_GM, "Margen bruto", "%",
                  hist=lambda y, gl, c: "=IFERROR(%s%d/%s%d,\"\")" % (gl, CG_GP, gl, CG_REV),
                  proj=lambda y, gl, c: "=%s%d+(%s-%s%d)*%d/%d" % (
                      lh, CG_GM, self.A("Margen bruto de salida"), lh, CG_GM,
                      self.proj_years.index(y) + 1, n),
                  fmt=PCT, bold=True)
        self.line(ws, CG_GP, "Beneficio bruto", "M", cd.hist.get("gross_profit", {}),
                  proj=lambda y, gl, c: "=%s%d*%s%d" % (gl, CG_REV, gl, CG_GM))
        self.line(ws, CG_COGS, "Coste de ventas", "M",
                  hist=lambda y, gl, c: "=-(%s%d-%s%d)" % (gl, CG_REV, gl, CG_GP),
                  proj=lambda y, gl, c: "=-(%s%d-%s%d)" % (gl, CG_REV, gl, CG_GP))
        return ws

    # --- PL ------------------------------------------------------------

    def sh_pl(self):
        cd = self.cd
        ws = self.wb.create_sheet("PL")
        _cell(ws, 1, 1, "%s — Cuenta de resultados" % cd.name, font=TITLE)
        self.year_header(ws, 3)
        lh, n = self.lh, self.n_proj

        self.line(ws, PL_REV, "Ingresos", "M",
                  hist=lambda y, gl, c: "=Revenue_Build!%s%d" % (gl, RB_REV),
                  proj=lambda y, gl, c: "=Revenue_Build!%s%d" % (gl, RB_REV), bold=True)
        self.line(ws, PL_GP, "Beneficio bruto", "M",
                  hist=lambda y, gl, c: "=COGS_GP!%s%d" % (gl, CG_GP),
                  proj=lambda y, gl, c: "=COGS_GP!%s%d" % (gl, CG_GP))
        self.line(ws, PL_OPEX, "Gastos operativos (SG&A + I+D + otros)", "M",
                  hist=lambda y, gl, c: "=IFERROR(%s%d-%s%d,\"\")" % (gl, PL_EBIT, gl, PL_GP),
                  proj=lambda y, gl, c: "=-%s%d*(ABS(%s%d)/%s%d+(%s-ABS(%s%d)/%s%d)*%d/%d)" % (
                      gl, PL_REV, lh, PL_OPEX, lh, PL_REV,
                      self.A("Opex de salida (% ingresos)"), lh, PL_OPEX, lh, PL_REV,
                      self.proj_years.index(y) + 1, n))
        self.line(ws, PL_EBIT, "EBIT (resultado de explotacion)", "M", cd.hist.get("ebit", {}),
                  proj=lambda y, gl, c: "=%s%d+%s%d" % (gl, PL_GP, gl, PL_OPEX), bold=True)
        self.line(ws, PL_MEBIT, "  Margen EBIT", "%",
                  hist=lambda y, gl, c: "=IFERROR(%s%d/%s%d,\"\")" % (gl, PL_EBIT, gl, PL_REV),
                  proj=lambda y, gl, c: "=IFERROR(%s%d/%s%d,\"\")" % (gl, PL_EBIT, gl, PL_REV),
                  fmt=PCT)
        self.line(ws, PL_DA, "Amortizaciones y depreciaciones", "M", cd.hist.get("da", {}),
                  proj=lambda y, gl, c: "=%s%d*%s" % (gl, PL_REV, self.A("D&A (% ingresos)")))
        self.line(ws, PL_EBITDA, "EBITDA", "M",
                  hist=lambda y, gl, c: "=IFERROR(%s%d+%s%d,\"\")" % (gl, PL_EBIT, gl, PL_DA),
                  proj=lambda y, gl, c: "=%s%d+%s%d" % (gl, PL_EBIT, gl, PL_DA), bold=True)
        self.line(ws, PL_MEBITDA, "  Margen EBITDA", "%",
                  hist=lambda y, gl, c: "=IFERROR(%s%d/%s%d,\"\")" % (gl, PL_EBITDA, gl, PL_REV),
                  proj=lambda y, gl, c: "=IFERROR(%s%d/%s%d,\"\")" % (gl, PL_EBITDA, gl, PL_REV),
                  fmt=PCT)
        self.line(ws, PL_INT, "Gastos financieros netos", "M", self.neg("interest"),
                  proj=lambda y, gl, c: "=-Balance!%s%d*%s" % (
                      gl, BS_DEBT, self.A("Coste medio de la deuda")))
        self.line(ws, PL_PBT, "Beneficio antes de impuestos", "M",
                  hist=lambda y, gl, c: "=IFERROR(%s%d+%s%d,\"\")" % (gl, PL_EBIT, gl, PL_INT),
                  proj=lambda y, gl, c: "=%s%d+%s%d" % (gl, PL_EBIT, gl, PL_INT))
        self.line(ws, PL_TAX, "Impuestos", "M", self.neg("tax"),
                  proj=lambda y, gl, c: "=-MAX(0,%s%d)*%s" % (
                      gl, PL_PBT, self.A("Tasa fiscal efectiva")))
        self.line(ws, PL_NI, "Beneficio neto", "M", cd.hist.get("net_income", {}),
                  proj=lambda y, gl, c: "=%s%d+%s%d" % (gl, PL_PBT, gl, PL_TAX), bold=True)
        _cell(ws, PL_NI + 2, 1, "Aviso: en negocios con mucha amortizacion de intangibles "
                                "adquiridos el beneficio neto GAAP no mide nada. Mira Cash_Flow.",
              font=SMALL)
        return ws

    # --- Working_Capital -----------------------------------------------

    def sh_wc(self):
        cd = self.cd
        ws = self.wb.create_sheet("Working_Capital")
        _cell(ws, 1, 1, "%s — Circulante" % cd.name, font=TITLE)
        _cell(ws, 2, 1, "Circulante NEGATIVO = el cliente financia el negocio: crecer genera "
                        "caja en vez de consumirla.", font=SMALL)
        self.year_header(ws, 4)

        self.line(ws, WC_REC, "Clientes", "M", cd.hist.get("receivables", {}))
        self.line(ws, WC_INV, "Existencias", "M", cd.hist.get("inventory", {}))
        self.line(ws, WC_PAY, "Proveedores", "M", cd.hist.get("payables", {}))
        self.line(ws, WC_NET, "Circulante neto (clientes + exist. - proveedores)", "M",
                  hist=lambda y, gl, c: "=IFERROR(%s%d+%s%d-%s%d,\"\")" % (
                      gl, WC_REC, gl, WC_INV, gl, WC_PAY),
                  proj=lambda y, gl, c: "=PL!%s%d*%s" % (
                      gl, PL_REV, self.A("Circulante neto (% ingresos)")), bold=True)
        self.line(ws, WC_PCT, "  Circulante / ingresos", "%",
                  hist=lambda y, gl, c: "=IFERROR(%s%d/PL!%s%d,\"\")" % (gl, WC_NET, gl, PL_REV),
                  proj=lambda y, gl, c: "=IFERROR(%s%d/PL!%s%d,\"\")" % (gl, WC_NET, gl, PL_REV),
                  fmt=PCT)
        self.line(ws, WC_DELTA, "Variacion del circulante (negativo = consume caja)", "M",
                  hist=lambda y, gl, c: ("=IFERROR(-(%s%d-%s%d),\"\")" % (
                      gl, WC_NET, get_column_letter(c - 1), WC_NET)) if c > C0 else None,
                  proj=lambda y, gl, c: "=-(%s%d-%s%d)" % (
                      gl, WC_NET, get_column_letter(c - 1), WC_NET), bold=True)
        _cell(ws, WC_CHK - 1, 1, "Comprobacion contra el estado de flujos reportado", font=BOLD)
        self.line(ws, WC_CHK, "  Var. circulante reportada", "M", cd.hist.get("d_wc", {}))
        return ws

    # --- Cash_Flow ------------------------------------------------------

    def sh_cf(self):
        cd = self.cd
        ws = self.wb.create_sheet("Cash_Flow")
        _cell(ws, 1, 1, "%s — Flujo de caja" % cd.name, font=TITLE)
        self.year_header(ws, 3)

        self.line(ws, CF_EBITDA, "EBITDA", "M",
                  hist=lambda y, gl, c: "=PL!%s%d" % (gl, PL_EBITDA),
                  proj=lambda y, gl, c: "=PL!%s%d" % (gl, PL_EBITDA))
        self.line(ws, CF_INT, "Intereses", "M",
                  hist=lambda y, gl, c: "=PL!%s%d" % (gl, PL_INT),
                  proj=lambda y, gl, c: "=PL!%s%d" % (gl, PL_INT))
        self.line(ws, CF_TAX, "Impuestos", "M",
                  hist=lambda y, gl, c: "=PL!%s%d" % (gl, PL_TAX),
                  proj=lambda y, gl, c: "=PL!%s%d" % (gl, PL_TAX))
        self.line(ws, CF_DWC, "Variacion del circulante", "M",
                  hist=lambda y, gl, c: "=IFERROR(Working_Capital!%s%d,0)" % (gl, WC_DELTA),
                  proj=lambda y, gl, c: "=Working_Capital!%s%d" % (gl, WC_DELTA))
        self.line(ws, CF_OTHER, "Otros (no monetarios, provisiones, SBC)", "M",
                  hist=lambda y, gl, c: "=IFERROR(%s%d-(%s%d+%s%d+%s%d+%s%d),0)" % (
                      gl, CF_OCF, gl, CF_EBITDA, gl, CF_INT, gl, CF_TAX, gl, CF_DWC),
                  proj=lambda y, gl, c: 0)
        self.line(ws, CF_OCF, "Flujo de caja de explotacion", "M", cd.hist.get("ocf", {}),
                  proj=lambda y, gl, c: "=%s%d+%s%d+%s%d+%s%d+%s%d" % (
                      gl, CF_EBITDA, gl, CF_INT, gl, CF_TAX, gl, CF_DWC, gl, CF_OTHER),
                  bold=True)
        self.line(ws, CF_CAPEX, "Capex", "M", self.neg("capex"),
                  proj=lambda y, gl, c: "=-PL!%s%d*%s" % (gl, PL_REV, self.A("Capex (% ingresos)")))
        self.line(ws, CF_FCF, "FREE CASH FLOW", "M",
                  hist=lambda y, gl, c: "=IFERROR(%s%d+%s%d,\"\")" % (gl, CF_OCF, gl, CF_CAPEX),
                  proj=lambda y, gl, c: "=%s%d+%s%d" % (gl, CF_OCF, gl, CF_CAPEX),
                  bold=True, fill_row=FILL_OK)
        self.line(ws, CF_MARGIN, "  Margen FCF", "%",
                  hist=lambda y, gl, c: "=IFERROR(%s%d/PL!%s%d,\"\")" % (gl, CF_FCF, gl, PL_REV),
                  proj=lambda y, gl, c: "=IFERROR(%s%d/PL!%s%d,\"\")" % (gl, CF_FCF, gl, PL_REV),
                  fmt=PCT)
        self.line(ws, CF_SBC, "Retribucion en acciones (SBC)", "M", cd.hist.get("sbc", {}),
                  proj=lambda y, gl, c: "=PL!%s%d*%s" % (gl, PL_REV, self.A("SBC (% ingresos)")))
        self.line(ws, CF_FCF_SBC, "FCF menos SBC (owner earnings)", "M",
                  hist=lambda y, gl, c: "=IFERROR(%s%d-%s%d,\"\")" % (gl, CF_FCF, gl, CF_SBC),
                  proj=lambda y, gl, c: "=%s%d-%s%d" % (gl, CF_FCF, gl, CF_SBC), bold=True)
        _cell(ws, CF_FCF_SBC + 1, 1,
              "  REGLA: el SBC SIEMPRE se resta. Es un gasto, no un apunte no monetario. "
              "La dilucion se mira en retrospectiva, no se proyecta (seria contarlo dos veces).",
              font=SMALL)

        _section(ws, CF_NORM_HDR, "FCF NORMALIZADO — solo para empresas que SOBREINVIERTEN",
                 len(self.years) + 2)
        self.line(ws, CF_MAINT_CAPEX, "Capex de mantenimiento", "M",
                  hist=lambda y, gl, c: "=-PL!%s%d*%s" % (
                      gl, PL_REV, self.A("Capex de MANTENIMIENTO (% ingresos)")),
                  proj=lambda y, gl, c: "=-PL!%s%d*%s" % (
                      gl, PL_REV, self.A("Capex de MANTENIMIENTO (% ingresos)")))
        self.line(ws, CF_GROWTH_CAPEX, "Capex de CRECIMIENTO (se devuelve)", "M",
                  hist=lambda y, gl, c: "=MAX(0,%s%d-%s%d)" % (
                      gl, CF_MAINT_CAPEX, gl, CF_CAPEX),
                  proj=lambda y, gl, c: "=MAX(0,%s%d-%s%d)" % (
                      gl, CF_MAINT_CAPEX, gl, CF_CAPEX))
        self.line(ws, CF_RND, "I+D reportada", "M", self.cd.hist.get("rnd", {}),
                  proj=lambda y, gl, c: None)
        self.line(ws, CF_RND_STEADY, "  de eso, I+D de mantenimiento", "M",
                  hist=lambda y, gl, c: "=IF(%s%d=\"\",0,PL!%s%d*%s)" % (
                      gl, CF_RND, gl, PL_REV, self.A("I+D de MANTENIMIENTO (% ingresos)")),
                  proj=lambda y, gl, c: "=PL!%s%d*%s" % (
                      gl, PL_REV, self.A("I+D de MANTENIMIENTO (% ingresos)")))
        self.line(ws, CF_RND_GROWTH, "  I+D de CRECIMIENTO (se devuelve)", "M",
                  hist=lambda y, gl, c: "=IFERROR(MAX(0,%s%d-%s%d),0)" % (
                      gl, CF_RND, gl, CF_RND_STEADY),
                  proj=lambda y, gl, c: "=MAX(0,%s%d-%s%d)" % (
                      gl, CF_RND, gl, CF_RND_STEADY))
        self.line(ws, CF_FCF_NORM, "FCF NORMALIZADO (tras SBC, sin inversion de crecimiento)",
                  "M",
                  hist=lambda y, gl, c: "=IFERROR(%s%d+%s%d+%s%d,\"\")" % (
                      gl, CF_FCF_SBC, gl, CF_GROWTH_CAPEX, gl, CF_RND_GROWTH),
                  proj=lambda y, gl, c: "=%s%d+%s%d+%s%d" % (
                      gl, CF_FCF_SBC, gl, CF_GROWTH_CAPEX, gl, CF_RND_GROWTH),
                  bold=True, fill_row=FILL_KEY)
        _cell(ws, CF_NORM_WARN, 1,
              "AVISO: si normalizas el FCF al alza, el multiplo de salida tiene que BAJAR. Un "
              "negocio que deja de invertir deja de crecer, y un negocio que no crece no vale "
              "20x. Normalizar Y subir el multiplo es contar el mismo argumento dos veces.",
              font=SMALL)
        _cell(ws, CF_NORM_WARN + 1, 1,
              "Rellena los dos supuestos de MANTENIMIENTO en Assumptions. Si los dejas a cero, "
              "esta seccion devuelve el FCF sin normalizar y no molesta.", font=SMALL)
        return ws

    # --- Balance --------------------------------------------------------

    def sh_balance(self):
        cd = self.cd
        ws = self.wb.create_sheet("Balance")
        _cell(ws, 1, 1, "%s — Balance (resumen)" % cd.name, font=TITLE)
        self.year_header(ws, 3)

        self.line(ws, BS_CASH, "Caja y equivalentes", "M", cd.hist.get("cash", {}),
                  proj=lambda y, gl, c: "=%s%d+Capital_Alloc!%s%d" % (
                      get_column_letter(c - 1), BS_CASH, gl, CA_RET))
        self.line(ws, BS_DEBT, "Deuda total", "M", cd.hist.get("debt", {}),
                  proj=lambda y, gl, c: "=%s%d" % (get_column_letter(c - 1), BS_DEBT))
        self.line(ws, BS_ND, "Deuda neta", "M",
                  hist=lambda y, gl, c: "=IFERROR(%s%d-%s%d,\"\")" % (gl, BS_DEBT, gl, BS_CASH),
                  proj=lambda y, gl, c: "=%s%d-%s%d" % (gl, BS_DEBT, gl, BS_CASH), bold=True)
        self.line(ws, BS_LEV, "  Deuda neta / EBITDA", "x",
                  hist=lambda y, gl, c: "=IFERROR(%s%d/PL!%s%d,\"\")" % (gl, BS_ND, gl, PL_EBITDA),
                  proj=lambda y, gl, c: "=IFERROR(%s%d/PL!%s%d,\"\")" % (gl, BS_ND, gl, PL_EBITDA),
                  fmt=MULT)
        self.line(ws, BS_EQ, "Fondos propios", "M", cd.hist.get("equity", {}),
                  proj=lambda y, gl, c: "=%s%d+PL!%s%d+Capital_Alloc!%s%d" % (
                      get_column_letter(c - 1), BS_EQ, gl, PL_NI, gl, CA_SHRET))
        self.line(ws, BS_GW, "Fondo de comercio", "M", cd.hist.get("goodwill", {}))
        self.line(ws, BS_ASSETS, "Activo total", "M", cd.hist.get("assets", {}))
        self.line(ws, BS_CE, "Capital empleado (FP + deuda neta)", "M",
                  hist=lambda y, gl, c: "=IFERROR(%s%d+%s%d,\"\")" % (gl, BS_EQ, gl, BS_ND),
                  proj=lambda y, gl, c: "=%s%d+%s%d" % (gl, BS_EQ, gl, BS_ND))
        self.line(ws, BS_ROIC, "ROIC (NOPAT / capital empleado)", "%",
                  hist=lambda y, gl, c: "=IFERROR(PL!%s%d*(1-%s)/%s%d,\"\")" % (
                      gl, PL_EBIT, self.A("Tasa fiscal efectiva"), gl, BS_CE),
                  proj=lambda y, gl, c: "=IFERROR(PL!%s%d*(1-%s)/%s%d,\"\")" % (
                      gl, PL_EBIT, self.A("Tasa fiscal efectiva"), gl, BS_CE),
                  fmt=PCT, bold=True)
        return ws

    # --- Capital_Alloc ---------------------------------------------------

    def sh_capalloc(self):
        cd = self.cd
        ws = self.wb.create_sheet("Capital_Alloc")
        _cell(ws, 1, 1, "%s — Asignacion de capital" % cd.name, font=TITLE)
        _cell(ws, 2, 1, "A donde va cada euro de free cash flow. En un compounder esta hoja "
                        "es la tesis entera.", font=SMALL)
        self.year_header(ws, 4)

        self.line(ws, CA_FCF, "Free cash flow", "M",
                  hist=lambda y, gl, c: "=Cash_Flow!%s%d" % (gl, CF_FCF),
                  proj=lambda y, gl, c: "=Cash_Flow!%s%d" % (gl, CF_FCF), bold=True)
        self.line(ws, CA_MA, "Adquisiciones (M&A neto)", "M", self.neg("ma"),
                  proj=lambda y, gl, c: "=-%s%d*%s" % (gl, CA_FCF, self.A("M&A (% del FCF)")))
        self.line(ws, CA_DIV, "Dividendos", "M", self.neg("dividends"),
                  proj=lambda y, gl, c: "=-%s%d*%s" % (gl, CA_FCF, self.A("Dividendos (% del FCF)")))
        self.line(ws, CA_BB, "Recompras netas de acciones", "M", self.neg("buybacks"),
                  proj=lambda y, gl, c: "=-%s%d*%s" % (gl, CA_FCF, self.A("Recompras (% del FCF)")))
        self.line(ws, CA_RET, "Caja retenida (va al balance)", "M",
                  hist=lambda y, gl, c: "=IFERROR(%s%d+%s%d+%s%d+%s%d,\"\")" % (
                      gl, CA_FCF, gl, CA_MA, gl, CA_DIV, gl, CA_BB),
                  proj=lambda y, gl, c: "=%s%d+%s%d+%s%d+%s%d" % (
                      gl, CA_FCF, gl, CA_MA, gl, CA_DIV, gl, CA_BB), bold=True)
        self.line(ws, CA_SHRET, "Retorno al accionista (div + recompras)", "M",
                  hist=lambda y, gl, c: "=IFERROR(%s%d+%s%d,\"\")" % (gl, CA_DIV, gl, CA_BB),
                  proj=lambda y, gl, c: "=%s%d+%s%d" % (gl, CA_DIV, gl, CA_BB))
        _cell(ws, CA_ROIIC - 1, 1, "Retorno del capital reinvertido en M&A", font=BOLD)
        self.line(ws, CA_ROIIC, "  Aumento de EBIT / M&A del anio anterior", "%",
                  hist=lambda y, gl, c: ("=IFERROR((PL!%s%d-PL!%s%d)/ABS(%s%d),\"\")" % (
                      gl, PL_EBIT, get_column_letter(c - 1), PL_EBIT,
                      get_column_letter(c - 1), CA_MA)) if c > C0 else None,
                  proj=lambda y, gl, c: "=IFERROR((PL!%s%d-PL!%s%d)/ABS(%s%d),\"\")" % (
                      gl, PL_EBIT, get_column_letter(c - 1), PL_EBIT,
                      get_column_letter(c - 1), CA_MA), fmt=PCT)
        return ws

    # --- Valuation -------------------------------------------------------

    def sh_valuation(self):
        cd = self.cd
        ws = self.wb.create_sheet("Valuation")
        for c, w in zip("ABCDE", (50, 6, 16, 16, 16)):
            ws.column_dimensions[c].width = w
        _cell(ws, 1, 1, "%s — Valoracion por escenarios" % cd.name, font=TITLE)
        _cell(ws, 2, 1, "Los tres escenarios se calculan a la vez: NO dependen de "
                        "Assumptions!B4. Horizonte 5 anios en los tres.", font=SMALL)
        _section(ws, 4, "VALORACION A 5 ANIOS (%d)" % self.term, 5)
        for j, t in enumerate(("Conservador", "Base", "Alcista")):
            _cell(ws, 4, 3 + j, t, font=H1, fill=FILL_HDR, align="center")

        S = lambda s: "CDE"[s - 1]
        tc = self.tc
        rows = [
            (V_REV, "Ingresos %d" % self.term, NUM, False,
             lambda s: "=Revenue_Build!$%s$%d*((1+%s+%s)/(1+%s+%s))^5" % (
                 tc, RB_REV, self.AS("Crecimiento organico anual", s),
                 self.AS("Crecimiento por adquisiciones anual", s),
                 self.A("Crecimiento organico anual"),
                 self.A("Crecimiento por adquisiciones anual"))),
            (V_MFCF, "Margen FCF %d" % self.term, PCT, False,
             lambda s: "=%s" % self.AS("Margen FCF de salida", s)),
            (V_FCF, "FCF %d antes de SBC" % self.term, NUM, False,
             lambda s: "=%s%d*%s%d" % (S(s), V_REV, S(s), V_MFCF)),
            (V_SBC, "menos SBC %d" % self.term, NUM, False,
             lambda s: "=-%s%d*%s" % (S(s), V_REV, self.AS("SBC (% ingresos)", s))),
            (V_FCFADJ, "FCF %d DESPUES DE SBC  (el que manda)" % self.term, NUM, True,
             lambda s: "=%s%d+%s%d" % (S(s), V_FCF, S(s), V_SBC)),
            (V_FCF_XL, "  (referencia) FCF@5y del Excel, antes de SBC", NUM, False, None),
            (V_EVFCF5, "EV DE HOY / FCF@5Y   <-- a que multiplo del flujo de %d compras"
             % self.term, MULT, True,
             lambda s: "=IFERROR(Assumptions!$B$%d/%s%d,\"n.s.\")" % (A_EV, S(s), V_FCFADJ)),
            (V_MULT, "Multiplo de salida EV/FCF", MULT, False,
             lambda s: "=%s" % self.AS("Multiplo de salida EV/FCF", s)),
            (V_EV, "Enterprise value %d" % self.term, NUM, False,
             lambda s: "=%s%d*%s%d" % (S(s), V_FCFADJ, S(s), V_MULT)),
            (V_ND, "Deuda neta usada (la de HOY, como en el Excel)", NUM, False,
             lambda s: "=Assumptions!$B$%d" % A_ND),
            (V_EQ, "Valor del equity %d" % self.term, NUM, False,
             lambda s: "=%s%d-%s%d" % (S(s), V_EV, S(s), V_ND)),
            (V_SH, "Acciones (M)", NUM1, False, lambda s: "=Assumptions!$B$%d" % A_SHARES),
            (V_PS, "VALOR POR ACCION %d (en moneda de cotizacion)" % self.term, NUM2, True,
             lambda s: "=%s%d/%s%d*Assumptions!$B$%d" % (S(s), V_EQ, S(s), V_SH, A_FX)),
            (V_PX, "Precio hoy", NUM2, False, lambda s: "=Assumptions!$B$%d" % A_PRICE),
            (V_UP, "Revalorizacion total", PCT, False,
             lambda s: "=IFERROR(%s%d/%s%d-1,\"\")" % (S(s), V_PS, S(s), V_PX)),
            (V_IRR, "TIR ANUALIZADA (5 anios)", PCT, True,
             lambda s: "=IFERROR((%s%d/%s%d)^(1/5)-1,\"\")" % (S(s), V_PS, S(s), V_PX)),
        ]
        xl_target = {1: cd.wl.get("fcf_5y_min"), 2: None, 3: cd.wl.get("fcf_5y_max")}
        for row, label, fmt, bold, fn in rows:
            _cell(ws, row, 1, label, font=BOLD if bold else BLACK)
            for s in (1, 2, 3):
                if fn is None:
                    _cell(ws, row, 2 + s, xl_target.get(s), font=BLUE, fmt=fmt)
                else:
                    _cell(ws, row, 2 + s, fn(s), font=BOLD if bold else BLACK, fmt=fmt,
                          fill=FILL_KEY if row == V_EVFCF5 else (FILL_OK if bold else None),
                          border=bold)
        _cell(ws, V_EVFCF5, 6, "Es la respuesta a: al precio de hoy, ¿cuantas veces el flujo "
                               "de caja de %d estoy pagando?" % self.term, font=SMALL)

        _cell(ws, V_ASYM - 1, 1, "ASIMETRIA", font=BOLD)
        # MISMA definicion que src/analytics.py::_irr_asymmetry_ratio, para que el modelo
        # y el dashboard cuantitativo no den numeros distintos:
        #     (TIR_alcista - TIR_bajista) / MAX(|TIR_bajista|, 5%)
        # El suelo del 5% evita que una TIR bajista cercana a cero infle el ratio.
        _cell(ws, V_ASYM, 1, "(TIR alcista - TIR bajista) / MAX(|TIR bajista|; 5%)", font=BLACK)
        _cell(ws, V_ASYM, 3, "=IFERROR((E%d-C%d)/MAX(ABS(C%d),0.05),\"n/d\")"
              % (V_IRR, V_IRR, V_IRR), font=BOLD, fmt=RATIO, fill=FILL_OK, border=True)
        _cell(ws, V_ASYM, 5, "=IF(ABS(C%d)<0.05,\"OJO: el suelo del 5%% manda. El ratio mide "
              "el diferencial, no el riesgo real de perdida\",\"\")" % V_IRR, font=SMALL)
        _cell(ws, V_MIN, 1, "Minimo exigido por el marco", font=BLACK)
        _cell(ws, V_MIN, 3, 2.0, font=BLUE, fmt=RATIO)
        _cell(ws, V_VERD, 1, "Veredicto", font=BOLD)
        _cell(ws, V_VERD, 3, "=IF(C%d>=C%d,\"PASA\",\"NO PASA\")" % (V_ASYM, V_MIN), font=BOLD)

        _cell(ws, V_BEARCHK - 1, 1, "CONTROL DEL CASO BAJISTA", font=BOLD)
        _cell(ws, V_BEARCHK, 1, "El FCF conservador de %d, ¿crece respecto al LTM?" % self.term,
              font=BLACK)
        _cell(ws, V_BEARCHK, 3,
              "=IF(C%d>Assumptions!$B$%d,\"SI — hay que JUSTIFICARLO por escrito en la tesis\","
              "\"No: el bajista decrece\")" % (V_FCF, A_FCF_LTM), font=BOLD)
        _cell(ws, V_BEARCHK, 5,
              "No es un error de por si. Un bajista puede crecer y aun asi perder dinero si las "
              "expectativas son altas (Prysmian: cartera de pedidos y picos y palas de IA, pero "
              "con momentum ya en el precio). Lo que NO vale es no haberlo pensado.", font=SMALL)
        _cell(ws, V_BEARCHK + 1, 1, "Nota: la TIR no incluye dividendos cobrados por el camino.",
              font=SMALL)

        _cell(ws, V_BEARCHK + 2, 1,
              "Deuda neta %d segun el roll-forward del balance (informativa)" % self.term,
              font=BLACK)
        _cell(ws, V_BEARCHK + 2, 3, "=Balance!$%s$%d" % (self.tc, BS_ND), font=BLACK, fmt=NUM)
        _cell(ws, V_BEARCHK + 3, 1,
              "  La valoracion usa la deuda neta de HOY, que es lo que hace el Excel y lo que "
              "no explota en financieras.", font=SMALL)

        _section(ws, V_NORM_HDR,
                 "SI LA EMPRESA SOBREINVIERTE — valoracion sobre FCF normalizado", 5)
        _cell(ws, V_FCFNORM, 1,
              "FCF normalizado %d = FCF base tras SBC + capex e I+D de CRECIMIENTO devueltos"
              % self.term, font=BLACK)
        _cell(ws, V_FCFNORM, 3,
              "=D%d+Cash_Flow!$%s$%d+Cash_Flow!$%s$%d" % (
                  V_FCFADJ, self.tc, CF_GROWTH_CAPEX, self.tc, CF_RND_GROWTH),
              font=BOLD, fmt=NUM, fill=FILL_KEY, border=True)
        _cell(ws, V_EVFCFNORM, 1, "EV DE HOY / FCF NORMALIZADO @5Y", font=BOLD)
        _cell(ws, V_EVFCFNORM, 3,
              "=IFERROR(Assumptions!$B$%d/C%d,\"n.s.\")" % (A_EV, V_FCFNORM),
              font=BOLD, fmt=MULT, fill=FILL_OK, border=True)
        _cell(ws, V_EVFCFNORM, 5,
              "Comparalo con el EV/FCF@5Y base de la fila %d. Si son parecidos, la empresa no "
              "sobreinvierte y esta seccion no aporta nada." % V_EVFCF5, font=SMALL)
        _cell(ws, V_NORMCHK, 1,
              "Control: si has normalizado al alza, ¿has BAJADO el multiplo de salida?",
              font=BLACK)
        _cell(ws, V_NORMCHK, 3,
              "=IF(C%d>D%d*1.05,\"SI has normalizado (+\"&TEXT(C%d/D%d-1,\"0%%\")&\"). "
              "BAJA el multiplo de salida: un negocio que deja de invertir deja de crecer, y "
              "contar las dos cosas es contar el mismo argumento dos veces\","
              "\"Sin normalizar: el capex ya esta por debajo del D&A\")"
              % (V_FCFNORM, V_FCFADJ, V_FCFNORM, V_FCFADJ), font=BLACK)

        _section(ws, V_RECON_HDR,
                 "RECONCILIACION — las dos rutas de FCF del modelo deberian parecerse", 5)
        _cell(ws, V_RECON_OP, 1,
              "Ruta A: modelo operativo (ingresos -> margenes -> caja), escenario activo B4",
              font=BLACK)
        _cell(ws, V_RECON_OP, 3, "=Cash_Flow!$%s$%d" % (self.tc, CF_FCF_SBC), font=BLACK, fmt=NUM)
        _cell(ws, V_RECON_CAL, 1,
              "Ruta B: FCF@5y calibrado al Excel de la watchlist (escenario base)", font=BLACK)
        _cell(ws, V_RECON_CAL, 3, "=D%d" % V_FCFADJ, font=BLACK, fmt=NUM)
        _cell(ws, V_RECON_GAP, 1, "Diferencia", font=BOLD)
        _cell(ws, V_RECON_GAP, 3, "=IFERROR(C%d/C%d-1,\"\")" % (V_RECON_OP, V_RECON_CAL),
              font=BOLD, fmt=PCT, border=True)
        _cell(ws, V_RECON_GAP, 5,
              "Si pasa del +-25%, una de las dos esta mal calibrada. Suele ser que el margen "
              "de opex de salida o el crecimiento del escenario no cuadran con el FCF@5y que "
              "hay en el Excel. Ponlo en la tesis antes de fiarte de la valoracion.", font=SMALL)
        return ws

    # --- Sensitivities ----------------------------------------------------

    def sh_sensitivities(self):
        ws = self.wb.create_sheet("Sensitivities")
        ws.column_dimensions["A"].width = 26
        for c in "BCDEF":
            ws.column_dimensions[c].width = 14
        _cell(ws, 1, 1, "%s — Sensibilidad de la TIR a 5 anios" % self.cd.name, font=TITLE)
        _cell(ws, 2, 1, "Filas: multiplo de salida. Columnas: FCF terminal. "
                        "Todo lo demas, del caso base.", font=SMALL)
        facs = [0.6, 0.8, 1.0, 1.2, 1.4]
        _cell(ws, 4, 1, "Multiplo \\ FCF terminal", font=BOLD, fill=FILL_BAND, border=True)
        for j, f in enumerate(facs):
            _cell(ws, 4, 2 + j, "=Valuation!$D$%d*%s" % (V_FCFADJ, f), font=BOLD, fmt=NUM,
                  align="center", fill=FILL_BAND, border=True)
        for i, m in enumerate(facs):
            r = 5 + i
            _cell(ws, r, 1, "=Valuation!$D$%d*%s" % (V_MULT, m), font=BOLD, fmt=MULT,
                  fill=FILL_BAND, border=True)
            for j in range(len(facs)):
                col = get_column_letter(2 + j)
                cell = _cell(ws, r, 2 + j,
                             "=IFERROR(((%s$4*$A%d-Valuation!$D$%d)/Valuation!$D$%d"
                             "*Assumptions!$B$%d/Assumptions!$B$%d)^(1/5)-1,\"\")" % (
                                 col, r, V_ND, V_SH, A_FX, A_PRICE),
                             font=BLACK, fmt=PCT, align="center", border=True)
                if i == 2 and j == 2:
                    cell.fill = FILL_OK
        _cell(ws, 11, 1, "La celda sombreada es el caso base de la hoja Valuation.", font=SMALL)
        return ws

    # --- Sources -----------------------------------------------------------

    def sh_sources(self):
        cd = self.cd
        ws = self.wb.create_sheet("Sources")
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 86
        _cell(ws, 1, 1, "%s — Fuentes y trazabilidad" % cd.name, font=TITLE)
        rng = "%d-%d" % (cd.years[0], cd.years[-1]) if cd.years else "n/d"
        items = [
            ("Historico P&L / flujos / balance",
             "yfinance sobre %s (income_stmt, cashflow, balance_sheet anuales). "
             "Ejercicios %s. Millones de %s." % (cd.yf_symbol, rng, cd.currency)),
            ("LTM / TTM",
             "Suma de los 4 ultimos trimestres de quarterly_income_stmt y quarterly_cashflow."),
            ("Precio, acciones, deuda, caja",
             "docs/data/watchlist.json (generado desde watchlist_ratings.xlsx)."),
            ("Escenarios FCF@5y y multiplos de salida",
             "Columnas BD/BF/BH/BI de watchlist_ratings.xlsx. Los mantiene Roger a mano."),
            ("Ratings R1/R2/R3",
             "watchlist_ratings.xlsx. R1 estructural /16, R2 calidad economica /100, "
             "R3 durabilidad y riesgo existencial /70."),
            ("NIVEL DE ESTE LIBRO",
             "Generado automaticamente (tools/build_pack.py). El historico es real y "
             "reportado y las formulas estan conectadas, pero el desglose de ingresos por "
             "segmento, los comentarios de fuente celda a celda y la calibracion fina de los "
             "escenarios son trabajo manual pendiente. Prysmian y Robertet siguen siendo el "
             "nivel de referencia."),
        ]
        r = 3
        for k, v in items:
            _cell(ws, r, 1, k, font=BOLD)
            _cell(ws, r, 2, v, font=BLACK, wrap=True)
            ws.row_dimensions[r].height = 34
            r += 1
        r += 1
        _cell(ws, r, 1, "Notas de extraccion", font=BOLD)
        for n in cd.notes:
            r += 1
            _cell(ws, r, 2, n, font=SMALL)
        return ws

    # --- Dashboard ----------------------------------------------------------

    def sh_dashboard(self):
        cd, wl = self.cd, self.cd.wl
        ws = self.wb.create_sheet("Dashboard")
        for c, w in zip("ABCDEF", (44, 18, 16, 16, 16, 16)):
            ws.column_dimensions[c].width = w
        ws.sheet_view.showGridLines = False
        _cell(ws, 1, 1, "%s (%s)" % (cd.name, cd.ticker), font=TITLE)
        _cell(ws, 2, 1, "%s · %s" % (wl.get("category", ""), wl.get("style", "")), font=SMALL)

        r = _section(ws, 4, "EJE 1 — CALIDAD  (no depende del precio)", 6)
        for label, key, fmt in [
            ("Rating compuesto", "rating_composite", NUM2),
            ("  R1 estructural", "rating_1", NUM2),
            ("  R2 calidad economica", "rating_2", NUM2),
            ("  R3 durabilidad", "rating_3", NUM2),
            ("ROIC LTM", "roic", PCT),
            ("Riesgo terminal (mas negativo = peor)", "r3_terminal_risk", NUM),
        ]:
            _cell(ws, r, 1, label, font=BOLD if key == "rating_composite" else BLACK)
            _cell(ws, r, 2, wl.get(key), font=BLUE, fmt=fmt, border=True,
                  fill=FILL_OK if key == "rating_composite" else None)
            r += 1

        r = _section(ws, r + 1, "EJE 2 — ASIMETRIA  (depende del precio de hoy)", 6)
        for label, formula, fmt, bold in [
            ("Precio", "=Assumptions!B%d" % A_PRICE, NUM2, False),
            ("EV / FCF LTM (tal cual)", "=Assumptions!B%d" % A_EVFCF, MULT, False),
            ("EV/FCF@5Y bajista", "=Valuation!C%d" % V_EVFCF5, MULT, False),
            ("EV/FCF@5Y BASE  <-- lo barata que esta", "=Valuation!D%d" % V_EVFCF5, MULT, True),
            ("EV/FCF@5Y alcista", "=Valuation!E%d" % V_EVFCF5, MULT, False),
            ("TIR conservadora (5a)", "=Valuation!C%d" % V_IRR, PCT, False),
            ("TIR base (5a)", "=Valuation!D%d" % V_IRR, PCT, False),
            ("TIR alcista (5a)", "=Valuation!E%d" % V_IRR, PCT, False),
            ("ASIMETRIA", "=Valuation!C%d" % V_ASYM, RATIO, True),
            ("Bajista que crece?", "=Valuation!C%d" % V_BEARCHK, None, False),
            ("EV/FCF normalizado @5Y", "=Valuation!C%d" % V_EVFCFNORM, MULT, False),
        ]:
            _cell(ws, r, 1, label, font=BOLD if bold else BLACK)
            _cell(ws, r, 2, formula, font=BOLD if bold else BLACK, fmt=fmt, border=bold,
                  fill=FILL_OK if bold else None)
            r += 1

        r = _section(ws, r + 1, "VEREDICTO", 6)
        v = self.cfg.get("verdict", {})
        for label, key in [("Conviccion (1-5)", "conviction"), ("Riesgo terminal", "terminal"),
                           ("Sizing recomendado", "sizing"), ("Alerta de precio", "alert"),
                           ("Proximo catalizador", "catalyst")]:
            _cell(ws, r, 1, label, font=BOLD)
            _cell(ws, r, 2, v.get(key, "pendiente"), font=BLUE, border=True)
            r += 1

        r += 1
        _cell(ws, r, 1, "Como usar este libro", font=BOLD)
        for i, t in enumerate([
            "1. Actualiza Assumptions!B8 (precio) antes de mirar nada de valoracion.",
            "2. Cambia Assumptions!B4 (1/2/3) para mover el modelo operativo entero.",
            "3. Valuation muestra los tres escenarios a la vez; no depende de B4.",
            "4. El SBC SIEMPRE se resta: la valoracion ya usa 'FCF despues de SBC'.",
            "5. EV/FCF@5Y es la metrica de cabecera: a que multiplo del flujo del anio 5 compras hoy.",
            "6. Si la empresa sobreinvierte, rellena los dos supuestos de MANTENIMIENTO y mira",
            "   el bloque de FCF normalizado. Si normalizas al alza, BAJA el multiplo de salida.",
        ]):
            _cell(ws, r + 1 + i, 1, t, font=SMALL)
        return ws
